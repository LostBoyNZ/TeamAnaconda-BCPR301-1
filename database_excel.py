# Graham
import sys

from openpyxl import load_workbook
from datetime import datetime

try:
    from errors import ErrorHandler as err
except NameError and ModuleNotFoundError and ImportError:
    print("Fatal Error - errors.py not found.")
    sys.exit()

try:
    from data_processor import DataProcessor as dp
except NameError and ModuleNotFoundError and ImportError:
    print(err.get_error_message(404, "data_processor"))
    sys.exit()

try:
    from validators.validate_date import ValidateDate as vd
except NameError and ModuleNotFoundError and ImportError:
    print(err.get_error_message(404, "validate_date"))
    sys.exit()

try:
    from log_file_handler import LogFileHandler as lfh
except NameError and ModuleNotFoundError and ImportError:
    print(err.get_error_message(404, "log_file_handler"))
    sys.exit()

# from file_reader import FileReader as fr

# try:
#     from file_reader import FileReader as fr
# except NameError and ModuleNotFoundError and ImportError:
#     print(err.get_error_message(404, "file_reader"))
#     sys.exit()


class DatabaseExcel(object):  # Graham

    row_names = ['empid', 'gender', 'age', 'sales', 'bmi', 'salary', 'birthday']

    def convert_date_format(self, excel_date):
        # excel_date = 42139
        dt = datetime.fromordinal(datetime(1900, 1, 1).toordinal() + excel_date - 2)
        output = dt.strftime("%d/%m/%Y")

        return output

    def choose_sheet(self, wb):
        sheet_names = wb.get_sheet_names()
        # default to the first sheet
        sheet = wb.get_sheet_by_name(sheet_names[0])

        if len(sheet_names) > 1:
            print("Sheets:\n")
            for sheet in sheet_names:
                print(sheet)

            while True:
                sheet_name = input("\nPlease enter the sheet to read data from >>> ")
                try:
                    sheet = wb.get_sheet_by_name(sheet_name)
                    break
                except KeyError:
                    print(err.get_error_message(501, sheet_name))

        return sheet

    def choose_file(self):
        while True:
            file_name = input("\nPlease enter the excel file to read >>> ")

            try:
                wb = load_workbook(file_name)
                break
            except FileNotFoundError:
                print(err.get_error_message(201))
            except OSError:
                print(err.get_error_message(103))

        return wb

    def create_connection(self, wb, switch):

        sheet = self.choose_sheet(wb)

        target_column = 2
        target_row = 1

        max_column = sheet.max_column
        max_row = sheet.max_row

        data_row = []
        row_dict = {}
        keys = []
        data_to_process = {}

        i = 0
        dup_keys = 0
        for row in self.row_names:
            row_dict[self.row_names[i]]: ''
            i = i + 1

        for row in range(0, max_row):

            # Get the first value from the row to set as the key
            output = sheet.cell(row=target_row, column=1).value
            key = dp.validate_key(str(output))

            # Check if it's a duplicate key
            if key in keys:
                dup_keys += 1
                data_to_log = "Duplicate Key" + str(key)
                lfh.append_file('log.txt', data_to_log)

            # Add that key to the list of all keys
            keys.append(key)
            data_to_process[key] = {}

            col_num = 0
            for column in range(0, max_column):
                output = sheet.cell(row=target_row, column=target_column).value
                data_row.append(str(output))

                row_dict[self.row_names[col_num]] = data_row[col_num]
                target_column = target_column + 1
                col_num = col_num + 1

            data_to_process[key]['gender'] = row_dict['gender']
            data_to_process[key]['age'] = row_dict['age']
            data_to_process[key]['sales'] = row_dict['sales']
            data_to_process[key]['bmi'] = row_dict['bmi']
            data_to_process[key]['salary'] = row_dict['salary']
            data_to_process[key]['birthday'] = row_dict['birthday']

            data_row = []
            target_column = 1
            target_row = target_row + 1

        # Send the data to be processed
        dict_valid = dp.send_to_validate(data_to_process, switch, dup_keys)

        # Send the data to be saved into a file
        # i = fr()
        # i.write_file(dict_valid)

        return dict_valid