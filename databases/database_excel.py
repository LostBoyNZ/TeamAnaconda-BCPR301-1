# Graham
import csv
import sys

from openpyxl import load_workbook

try:
    from errors import ErrorHandler as err
except NameError and ModuleNotFoundError and ImportError:
    print("Fatal Error - errors.py not found.")
    sys.exit()


class ExcelReader(object):  # Graham

    def pick_sheet(self, wb):
        sheet_names = (wb.get_sheet_names())
        # default to the first sheet
        sheet_name = sheet_names[0]

        if len(sheet_names) > 1:
            print("Sheets:\n")
            for sheet in sheet_names:
                print(sheet)

            requested_sheet_name = input("\nPlease enter the sheet to read data from >>> ")
            sheet_name = requested_sheet_name

        try:
            sheet = wb.get_sheet_by_name(sheet_name)
        except KeyError:
            print(err.get_error_message(501, sheet_name))

        return sheet

    def start(self):

        file_name = '../test_data2.xlsx'

        wb = load_workbook(file_name)

        sheet = self.pick_sheet(wb)

        target_column = 1
        target_row = 1
        max_column = sheet.max_column
        max_row = sheet.max_row

        all_the_rows = []
        data_row = []

        for row in range(1, max_row):
            for column in range(1, max_column):
                output = sheet.cell(row=target_row, column=target_column).value
                data_row.append(output)
                target_column = target_column + 1
            all_the_rows.append(data_row)
            data_row = []
            target_column = 1
            target_row = target_row + 1

        print(all_the_rows)

i = ExcelReader()
i.start()
